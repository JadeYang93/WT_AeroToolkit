# -*- coding: utf-8 -*-
"""载荷预估核心算法。

源项目 F:/python/载荷信息读取/load_estimation.py 已剥离：
- 去除 BASE_DIR / DATA_DIR / OUTPUT_DIR 等硬编码路径常量
- save_results / plot_result 都接受外部传入的 output_dir / ax
- plot_result 抽出来供 UI 嵌入式画布复用（不再每次新建 figure）

不变量：
- 输入 Excel 必须含 3 个 sheet：baseLineSteady / baselineDynamic / newSteady
- baseLineSteady 列：[t, paramX, paramY, ...]
- baselineDynamic 列：[t, maxMx, minMx, maxMy, minMy]
- newSteady 列：[t, paramX, paramY, ...]

拟合关系：
- maxMx/minMx ← polyfit(baseLineSteady.paramX, baselineDynamic.maxMx/minMx, N_ORDER)
- maxMy/minMy ← polyfit(baseLineSteady.paramY, baselineDynamic.maxMy/minMy, N_ORDER)
- 新工况预测：把 newSteady.paramX/paramY 代入拟合多项式
"""
import os

import numpy as np
from openpyxl import load_workbook

# matplotlib 仅在 save_results / plot_result 内部用；
# 字体配置由调用方（CLI / UI）负责，这里不重复设置，避免 import 时机耦合。
import matplotlib.pyplot as plt


# 多项式拟合阶数（源项目默认 6）
N_ORDER = 6

# 4 个载荷分量
COMPONENTS = ['maxMx', 'minMx', 'maxMy', 'minMy']

# UI 查看下拉的 8 个选项（kind-name 形式，kind ∈ {baseline, new}）
VIEW_OPTIONS = [
    'baseline-maxMx', 'baseline-minMx', 'baseline-maxMy', 'baseline-minMy',
    'new-maxMx', 'new-minMx', 'new-maxMy', 'new-minMy',
]


def _sheet_to_array(ws) -> np.ndarray:
    """openpyxl worksheet → ndarray，跳过表头行。"""
    rows = list(ws.iter_rows(values_only=True))
    return np.array([[float(v) for v in row[1:]] for row in rows[1:]], dtype=float)


def load_data(xlsx_path):
    """从 Excel 读取三组数据。

    Returns:
        dict: {'base_steady', 'base_dynamic', 'new_steady'}，每个为 ndarray。
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        base_steady = _sheet_to_array(wb['baseLineSteady'])
        base_dynamic = _sheet_to_array(wb['baselineDynamic'])
        new_steady = _sheet_to_array(wb['newSteady'])
    finally:
        wb.close()

    # 行数断言：占位数据可能不等长，截断 dynamic 到与 steady 等长
    if base_steady.shape[0] != base_dynamic.shape[0]:
        print(
            f'[警告] baseLineSteady({base_steady.shape[0]} 行) 与 '
            f'baselineDynamic({base_dynamic.shape[0]} 行) 行数不等，'
            f'截断 dynamic 到前 {base_steady.shape[0]} 行。真实数据应保证等长。'
        )
        base_dynamic = base_dynamic[: base_steady.shape[0]]

    return {
        'base_steady': base_steady,
        'base_dynamic': base_dynamic,
        'new_steady': new_steady,
    }


def fit_loads(data, n=N_ORDER):
    """N 阶多项式拟合 + 新工况预测。

    Returns:
        dict 含 coeffs / baseline / new / result，结构见模块 docstring。
    """
    base_steady = data['base_steady']
    base_dynamic = data['base_dynamic']
    new_steady = data['new_steady']

    # 列约定（与源项目一致）：
    # base_steady[:, 0] = t, [:, 1] = paramX, [:, 2] = paramY
    # base_dynamic[:, 0..3] = maxMx, minMx, maxMy, minMy
    # new_steady[:, 1] = paramX, [:, 2] = paramY
    p_mx_max = np.polyfit(base_steady[:, 1], base_dynamic[:, 0], n)
    p_mx_min = np.polyfit(base_steady[:, 1], base_dynamic[:, 1], n)
    p_my_max = np.polyfit(base_steady[:, 2], base_dynamic[:, 2], n)
    p_my_min = np.polyfit(base_steady[:, 2], base_dynamic[:, 3], n)

    baseline_fits = {
        'maxMx': np.polyval(p_mx_max, base_steady[:, 1]),
        'minMx': np.polyval(p_mx_min, base_steady[:, 1]),
        'maxMy': np.polyval(p_my_max, base_steady[:, 2]),
        'minMy': np.polyval(p_my_min, base_steady[:, 2]),
    }
    new_preds = {
        'maxMx': np.polyval(p_mx_max, new_steady[:, 1]),
        'minMx': np.polyval(p_mx_min, new_steady[:, 1]),
        'maxMy': np.polyval(p_my_max, new_steady[:, 2]),
        'minMy': np.polyval(p_my_min, new_steady[:, 2]),
    }

    return {
        'coeffs': {
            'p_mx_max': p_mx_max, 'p_mx_min': p_mx_min,
            'p_my_max': p_my_max, 'p_my_min': p_my_min,
        },
        'baseline': {
            't': base_steady[:, 0],
            'original': {
                'maxMx': base_dynamic[:, 0], 'minMx': base_dynamic[:, 1],
                'maxMy': base_dynamic[:, 2], 'minMy': base_dynamic[:, 3],
            },
            'fitted': baseline_fits,
        },
        'new': {
            't': new_steady[:, 0],
            'pred': new_preds,
        },
        'result': np.column_stack([
            new_steady[:, 0],
            new_preds['maxMx'], new_preds['minMx'],
            new_preds['maxMy'], new_preds['minMy'],
        ]),
    }


def save_results(results, output_dir):
    """保存 result.csv / coefficients.csv + 8 张 PNG 到 output_dir。

    output_dir 由调用方决定（CLI 走默认 output/；工具箱走 ConfigCenter 配的输出目录）。
    """
    os.makedirs(output_dir, exist_ok=True)

    # result.csv：新工况预测结果
    np.savetxt(
        os.path.join(output_dir, 'result.csv'),
        results['result'], delimiter=',',
        header='zspan,maxMx,minMx,maxMy,minMy',
        comments='', fmt='%.6f',
    )

    # coefficients.csv：4 组多项式系数（polyfit 返回顺序：高次→低次）
    c = results['coeffs']
    coeffs = np.array([c['p_mx_max'], c['p_mx_min'], c['p_my_max'], c['p_my_min']])
    np.savetxt(
        os.path.join(output_dir, 'coefficients.csv'),
        coeffs, delimiter=',',
        header='c_x6,c_x5,c_x4,c_x3,c_x2,c_x1,c_x0',
        comments='', fmt='%.10e',
    )

    # 8 张 PNG（baseline 4 + new 4）
    t_base = results['baseline']['t']
    t_new = results['new']['t']
    for name in COMPONENTS:
        orig = results['baseline']['original'][name]
        fit = results['baseline']['fitted'][name]
        fig, ax = plt.subplots(figsize=(10, 5))
        ax.plot(t_base, orig, 'k', label='original', linewidth=1.5)
        ax.plot(t_base, fit, 'r', label='fitted', linewidth=2)
        ax.set_title(f'{name} fitting (baseline)')
        ax.set_xlabel('zspan')
        ax.set_ylabel(name)
        ax.legend()
        ax.grid(True)
        fig.tight_layout()
        fig.savefig(os.path.join(output_dir, f'figure_baseline_{name}.png'), dpi=150)
        plt.close(fig)

    for name in COMPONENTS:
        pred = results['new']['pred'][name]
        fig, ax = plt.subplots(figsize=(10, 5))
        ax.plot(t_new, pred, 'b', label=name, linewidth=2)
        ax.set_title(f'{name} prediction (new steady)')
        ax.set_xlabel('zspan')
        ax.set_ylabel(name)
        ax.legend()
        ax.grid(True)
        fig.tight_layout()
        fig.savefig(os.path.join(output_dir, f'figure_new_{name}.png'), dpi=150)
        plt.close(fig)


def plot_result(ax, results, kind: str, name: str):
    """在调用方提供的 matplotlib Axes 上画单张图（嵌入式 UI 复用）。

    kind ∈ {'baseline', 'new'}，name ∈ COMPONENTS。
    调用方负责 ax.clear() 与 canvas.draw()。
    """
    if kind == 'baseline':
        t = results['baseline']['t']
        orig = results['baseline']['original'][name]
        fit = results['baseline']['fitted'][name]
        ax.plot(t, orig, 'k', label='original', linewidth=1.5)
        ax.plot(t, fit, 'r', label='fitted', linewidth=2)
        ax.set_title(f'{name} fitting (baseline)')
    else:  # 'new'
        t = results['new']['t']
        pred = results['new']['pred'][name]
        ax.plot(t, pred, 'b', label=name, linewidth=2)
        ax.set_title(f'{name} prediction (new steady)')
    ax.set_xlabel('zspan')
    ax.set_ylabel(name)
    ax.legend()
    ax.grid(True)
