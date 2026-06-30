# -*- coding: utf-8 -*-
"""失速评估核心算法。

业务逻辑（无 PyQt5 依赖）。UI 面板见 ``src/tools/stall_assessment_panel.py``。

核心思路：
    由「相对厚度 → 失速攻角」的标准翼型表（离散对应关系）建立 PCHIP 保形插值，
    再把沿展向的相对厚度分布代入，得到失速攻角的展向分布。

不变量：
- 标准表与展向分布的相对厚度单位必须一致（同用小数或同用百分数），不做单位换算。
- 展向分布中若有相对厚度超出标准表范围 → 抛 ValueError（要求补全标准表，不外推）。
- 展向位置严格由输入决定，输出顺序与输入一致。

公共 API:
    parse_span_text(text)           → 文本粘贴 → (positions, thickness) ndarray
    parse_span_file(path)           → CSV/xlsx → (positions, thickness) ndarray
    interpolate(thickness, alpha,   → PCHIP 插值，返回展向失速攻角
                span_thickness)
    save_csv(positions, thickness,  → 写两列 CSV 到 output_dir
              alpha, out_dir)
    plot_check(ax, thickness,       → 校核图：标准点 + 插值曲线
               alpha, span_thickness, span_alpha)
    plot_span(ax, positions, alpha) → 展向分布图：r/R ↔ 失速攻角
"""
import os

import numpy as np
from scipy.interpolate import PchipInterpolator

# matplotlib 仅在画图函数内部用；字体配置由调用方（CLI / UI）负责。
import matplotlib.pyplot as plt


# ============================================================
# 输入解析
# ============================================================
def _rows_to_array(rows):
    """把「每行两列」的字符串行序列转成 (N, 2) ndarray。

    容错策略（与展向分布输入约定一致）：
    - 自动跳过空行；
    - 首行若不能解析为两个数字（即表头），整行跳过。
    - 列分隔符兼容逗号 / 制表符 / 连续空格。
    """
    parsed = []
    for line in rows:
        line = line.strip()
        if not line:
            continue
        # 统一分隔符：逗号 / 制表符 → 空格，再用 split 处理多空格
        parts = line.replace(',', ' ').replace('\t', ' ').split()
        if len(parts) < 2:
            continue
        try:
            a, b = float(parts[0]), float(parts[1])
        except ValueError:
            # 首行表头（如 "r/R, t/c"）→ 跳过
            continue
        parsed.append((a, b))
    if not parsed:
        raise ValueError('未解析到任何有效数据行（至少需要「展向位置, 相对厚度」两列）。')
    return np.array(parsed, dtype=float)


def parse_span_text(text):
    """粘贴文本 → (positions, thickness)。

    text: 多行文本，每行「展向位置, 相对厚度」，首行可为表头。
    Returns:
        (positions[N], thickness[N]) 两个一维 ndarray。
    """
    arr = _rows_to_array(text.splitlines())
    return arr[:, 0], arr[:, 1]


def parse_span_file(path):
    """CSV / xlsx → (positions, thickness)。

    约定：取前两列为 (展向位置, 相对厚度)，首行可为表头。
    """
    ext = os.path.splitext(path)[1].lower()
    if ext in ('.xlsx', '.xls'):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            rows = [
                (str(r[0]), str(r[1])) if r[0] is not None and r[1] is not None
                else ''
                for r in ws.iter_rows(values_only=True)
            ]
            # openpyxl 读出的数值已是 float，转成 "a b" 形式走同一解析管线
            lines = []
            for r in ws.iter_rows(values_only=True):
                if r[0] is None and r[1] is None:
                    lines.append('')
                else:
                    lines.append(f'{r[0]} {r[1]}')
        finally:
            wb.close()
        arr = _rows_to_array(lines)
        return arr[:, 0], arr[:, 1]
    else:
        # CSV / TXT
        with open(path, 'r', encoding='utf-8-sig', errors='replace') as f:
            arr = _rows_to_array(f.readlines())
        return arr[:, 0], arr[:, 1]


def normalize_positions(positions):
    """展向位置无量纲化：若最大值 > 1，判定为实际位置，整列除以最大值。

    约定：无量纲展向位置 r/R ∈ [0, 1]。若输入最大值超过 1（如实际米数 0~85），
    视为实际展向位置，自动归一化到 [0, 1]；否则原样返回。

    Returns:
        (normalized, did_normalize): 归一化后的数组 + 是否发生了归一化。
    """
    positions = np.asarray(positions, dtype=float)
    m = positions.max()
    if m > 1.0:
        return positions / m, True
    return positions, False


# ============================================================
# 插值
# ============================================================
def interpolate(thickness, alpha, span_thickness):
    """PCHIP 保形插值：标准表 (thickness→alpha) 代入展向厚度。

    Args:
        thickness (array): 标准翼型相对厚度（横轴）。
        alpha (array):     标准翼型失速攻角（纵轴，度）。
        span_thickness (array): 展向各站的相对厚度。

    Returns:
        span_alpha (array): 展向各站的失速攻角（度）。

    Raises:
        ValueError: 标准点不足 2 个、厚度有重复、或展向厚度超出标准表范围。
    """
    thickness = np.asarray(thickness, dtype=float)
    alpha = np.asarray(alpha, dtype=float)
    span_thickness = np.asarray(span_thickness, dtype=float)

    if thickness.size < 2:
        raise ValueError('标准翼型表至少需要 2 个点才能插值。')
    # PCHIP 要求 x 严格递增；若用户乱序输入则排序
    order = np.argsort(thickness)
    thickness = thickness[order]
    alpha = alpha[order]
    if np.any(np.diff(thickness) == 0):
        raise ValueError('标准翼型表存在重复的相对厚度值，无法插值。')

    lo, hi = thickness.min(), thickness.max()
    out_of_range = span_thickness[(span_thickness < lo) | (span_thickness > hi)]
    if out_of_range.size:
        raise ValueError(
            f'展向相对厚度 {out_of_range.min():.4f}~{out_of_range.max():.4f} '
            f'超出标准表范围 [{lo:.4f}, {hi:.4f}]，请补全标准翼型表。'
        )

    spline = PchipInterpolator(thickness, alpha)
    return spline(span_thickness)


# ============================================================
# VG 支持：按相对厚度区间两端点的「标/VG」身份做两点线性插值
# ============================================================
def find_thickness_interval(t_z, std_thickness):
    """找 t_z 落在哪两个相邻标准厚度之间。

    Args:
        t_z (float): 当前 z 的相对厚度。
        std_thickness (array-like): 标准厚度数组（不必有序，函数内排序）。
    Returns:
        (t_high, t_low): 两个相邻标准厚度，t_high >= t_z >= t_low。
    Raises:
        ValueError: t_z 超出标准表范围。
    """
    sorted_t = np.sort(np.asarray(std_thickness, dtype=float))[::-1]  # 降序
    if t_z > sorted_t[0] + 1e-9 or t_z < sorted_t[-1] - 1e-9:
        raise ValueError(
            f'相对厚度 {t_z:.4f} 超出标准表范围 '
            f'[{sorted_t[-1]:.4f}, {sorted_t[0]:.4f}]，请补全标准翼型表。'
        )
    for i in range(len(sorted_t) - 1):
        if sorted_t[i] >= t_z >= sorted_t[i + 1]:
            return float(sorted_t[i]), float(sorted_t[i + 1])
    return float(sorted_t[-2]), float(sorted_t[-1])


def pick_alpha(thickness, z, alpha_std_map, alpha_vg_map, vg_segments):
    """标准厚度 `thickness` 在 z 处的失速攻角：VG 安装范围内 → α_VG，否则 α_std。

    Args:
        thickness (float): 标准厚度（如 30、40）。
        z (float): 展向位置 r/R。
        alpha_std_map (dict): {thickness: alpha_std}。
        alpha_vg_map (dict): {thickness: alpha_vg}（VG 列非空的子集）。
        vg_segments (list[tuple]): [(thickness, z_start, z_end), ...]。
    Returns:
        float: 失速攻角。
    """
    # 用 1e-9 容差比较 z 边界，避免 np.linspace 浮点误差导致边界点漏判
    for seg_t, seg_zs, seg_ze in vg_segments:
        if (abs(seg_t - thickness) < 1e-9
                and (seg_zs - 1e-9) <= z <= (seg_ze + 1e-9)):
            return alpha_vg_map[thickness]
    return alpha_std_map[thickness]


def compute_alpha_span(positions, thickness, std_thickness, std_alpha_std,
                       std_alpha_vg=None, vg_segments=None):
    """带 VG 支持的失速攻角展向插值。

    核心逻辑：对每个 z，找其厚度所在的标准厚度区间 [t_a, t_b]，分别查 t_a、t_b 在
    z 处用标还是 VG 攻角（取决于 VG 安装范围），再按相对厚度做两点线性插值。

    VG 安装范围空 / VG 攻角表空 → 退化为原 PCHIP 逻辑（兼容路径）。

    Args:
        positions (array): 展向位置 r/R，长度 N。
        thickness (array): 每个展向位置的相对厚度，长度 N。
        std_thickness (array): 标准厚度数组，长度 M。
        std_alpha_std (array): 标准厚度的标失速攻角，长度 M。
        std_alpha_vg (dict or None): {thickness: alpha_vg}，VG 列非空的子集。
        vg_segments (list[tuple] or None): [(thickness, z_start, z_end), ...]。
    Returns:
        (alpha_arr, vg_active): 长度 N 的失速攻角数组 + 长度 N 的布尔数组
            (该 z 是否落在任何 VG 安装范围内，不论厚度)。
    """
    # 兼容路径：无 VG 配置 → 用原 PCHIP
    if (not vg_segments) or (not std_alpha_vg):
        alpha = interpolate(std_thickness, std_alpha_std, thickness)
        return alpha, np.zeros(len(positions), dtype=bool)

    positions = np.asarray(positions, dtype=float)
    thickness = np.asarray(thickness, dtype=float)
    alpha_arr = np.empty(len(positions), dtype=float)
    vg_active = np.zeros(len(positions), dtype=bool)

    alpha_std_map = {float(t): float(a)
                     for t, a in zip(std_thickness, std_alpha_std)}
    # 预处理 VG 段（确保 float）
    seg_list = [(float(t), float(zs), float(ze))
                for t, zs, ze in vg_segments]

    for i in range(len(positions)):
        z = positions[i]
        t_z = thickness[i]
        t_high, t_low = find_thickness_interval(t_z, std_thickness)
        alpha_a = pick_alpha(t_high, z, alpha_std_map, std_alpha_vg, seg_list)
        alpha_b = pick_alpha(t_low, z, alpha_std_map, std_alpha_vg, seg_list)
        if t_high != t_low:
            alpha_arr[i] = alpha_a + (alpha_b - alpha_a) * \
                (t_high - t_z) / (t_high - t_low)
        else:
            # 极端特例：厚度区间退化为一点（用户配置不当）→ 直接取 a 端
            alpha_arr[i] = alpha_a
        # vg_active：z 是否落在任何 VG 安装段内（不论厚度），同样用 1e-9 容差
        for seg_t, seg_zs, seg_ze in seg_list:
            if (seg_zs - 1e-9) <= z <= (seg_ze + 1e-9):
                vg_active[i] = True
                break

    return alpha_arr, vg_active


# ============================================================
# 输出
# ============================================================
def save_csv(positions, thickness, alpha, output_dir,
             vg_active=None, filename='stall_alpha_span.csv'):
    """写展向失速攻角分布到 CSV。

    默认三列：位置 / 相对厚度 / 失速攻角。
    若传入 vg_active（一维布尔/0-1 数组），追加 `vg_active` 列（1.0/0.0）。
    """
    os.makedirs(output_dir, exist_ok=True)
    out_path = os.path.join(output_dir, filename)
    if vg_active is None:
        data = np.column_stack([positions, thickness, alpha])
        header = 'span_position,relative_thickness,stall_alpha_deg'
    else:
        vg_col = np.asarray(vg_active).astype(float).reshape(-1, 1)
        data = np.column_stack([positions, thickness, alpha, vg_col])
        header = 'span_position,relative_thickness,stall_alpha_deg,vg_active'
    np.savetxt(out_path, data, delimiter=',', header=header,
               comments='', fmt='%.6f')
    return out_path


def plot_check(ax, thickness, alpha, span_thickness, span_alpha):
    """校核图：标准翼型点（散点）+ 插值曲线（连续）。

    横轴相对厚度，纵轴失速攻角。让用户直观判断 PCHIP 拟合是否合理。
    调用方负责 ax.clear() 与 canvas.draw()。
    """
    order = np.argsort(thickness)
    t_sorted = np.asarray(thickness)[order]
    a_sorted = np.asarray(alpha)[order]

    # 插值曲线（用标准表范围内的细密采样画连续曲线）
    t_dense = np.linspace(t_sorted.min(), t_sorted.max(), 200)
    spline = PchipInterpolator(t_sorted, a_sorted)
    a_dense = spline(t_dense)

    ax.plot(t_dense, a_dense, '-', color='#1f77b4', linewidth=2, label='PCHIP 插值')
    ax.scatter(thickness, alpha, marker='o', color='#d62728', zorder=5,
               s=45, label='标准翼型点', edgecolors='white', linewidths=0.8)
    ax.set_xlabel('相对厚度')
    ax.set_ylabel('失速攻角 (°)')
    ax.set_title('插值校核')
    ax.grid(True, alpha=0.3)
    ax.legend(loc='best')


def plot_span(ax, positions, alpha):
    """展向分布图：展向位置 r/R ↔ 失速攻角。

    调用方负责 ax.clear() 与 canvas.draw()。
    """
    order = np.argsort(positions)
    p_sorted = np.asarray(positions)[order]
    a_sorted = np.asarray(alpha)[order]

    ax.plot(p_sorted, a_sorted, '-o', color='#2ca02c', linewidth=2,
            markersize=4, label='失速攻角')
    ax.set_xlabel('展向位置 (r/R)')
    ax.set_ylabel('失速攻角 (°)')
    ax.set_title('失速攻角展向分布')
    ax.grid(True, alpha=0.3)
    ax.legend(loc='best')


def find_intersections(stall_pos, stall_alpha, aoa_pos, aoa):
    """求失速攻角曲线与实际攻角曲线的相交点。

    两条曲线的展向位置可能不同，先用 PCHIP 把两者都插值到统一细密坐标，
    再找差值符号反转点（线性插值近似交点横坐标）。

    Args:
        stall_pos (array): 失速攻角曲线的展向位置。
        stall_alpha (array): 失速攻角值。
        aoa_pos (array): 实际攻角曲线的展向位置。
        aoa (array): 实际攻角值。

    Returns:
        list[float]: 相交点的展向位置（r/R），升序。可能为空。
    """
    stall_pos = np.asarray(stall_pos, dtype=float)
    stall_alpha = np.asarray(stall_alpha, dtype=float)
    aoa_pos = np.asarray(aoa_pos, dtype=float)
    aoa = np.asarray(aoa, dtype=float)

    # 统一到两条曲线共同覆盖的展向区间，细密采样
    lo = max(stall_pos.min(), aoa_pos.min())
    hi = min(stall_pos.max(), aoa_pos.max())
    if hi <= lo:
        return []
    x_dense = np.linspace(lo, hi, 1000)
    f_stall = PchipInterpolator(np.sort(stall_pos),
                                stall_alpha[np.argsort(stall_pos)])(x_dense)
    f_aoa = PchipInterpolator(np.sort(aoa_pos),
                              aoa[np.argsort(aoa_pos)])(x_dense)

    diff = f_stall - f_aoa
    crossings = []
    for i in range(len(diff) - 1):
        if diff[i] == 0:
            crossings.append(x_dense[i])
        elif diff[i] * diff[i + 1] < 0:
            # 线性插值近似零点
            t = diff[i] / (diff[i] - diff[i + 1])
            crossings.append(x_dense[i] + t * (x_dense[i + 1] - x_dense[i]))
    return crossings


def plot_span_compare(ax, stall_pos, stall_alpha, aoa_pos, aoa,
                      crossings=None, span_pos=None, span_thickness=None,
                      std_thickness=None, show_thickness=True):
    """双曲线对比图：失速攻角 + 最大攻角（左 y 轴）+ 相对厚度（右 y 轴）。

    横轴展向位置 r/R。
    - 左轴攻角 (°)：失速攻角 + 最大攻角两条 PCHIP 平滑曲线，相交点标出。
    - 右轴相对厚度 (%)：show_thickness=True 时画 PCHIP 平滑曲线。
    调用方负责 ax.clear() 与 canvas.draw()。
    """
    # ===== 左轴：攻角 =====
    # 失速攻角曲线（PCHIP 平滑，无标记点）
    order_s = np.argsort(stall_pos)
    xs_s = np.asarray(stall_pos)[order_s]
    ys_s = np.asarray(stall_alpha)[order_s]
    f_stall = PchipInterpolator(xs_s, ys_s)
    x_dense_s = np.linspace(xs_s.min(), xs_s.max(), 300)
    ax.plot(x_dense_s, f_stall(x_dense_s),
            '-', color='#2ca02c', linewidth=2, alpha=0.5, label='失速攻角')

    # 最大攻角曲线（PCHIP 平滑，无标记点）
    order_a = np.argsort(aoa_pos)
    xs_a = np.asarray(aoa_pos)[order_a]
    ys_a = np.asarray(aoa)[order_a]
    f_aoa = PchipInterpolator(xs_a, ys_a)
    x_dense_a = np.linspace(xs_a.min(), xs_a.max(), 300)
    ax.plot(x_dense_a, f_aoa(x_dense_a),
            '-', color='#1f77b4', linewidth=2, alpha=0.5, label='最大攻角')

    # 攻角相交点（黑实心圆 + 垂直引线 + 可拖标注）
    if crossings:
        xs = np.array(crossings)
        ys = f_stall(xs)
        ax.scatter(xs, ys, marker='o', color='black', s=45, zorder=6,
                   facecolors='black', label='失速位置')
        for xv, yv in zip(xs, ys):
            ax.axvline(xv, color='black', linestyle=':', linewidth=0.8,
                       alpha=0.5, zorder=1)
            ax.annotate(f'r/R = {xv:.3f}', (xv, yv),
                        textcoords='offset points', xytext=(18, 22),
                        ha='left', fontsize=9, color='black',
                        fontweight='bold', picker=True,
                        arrowprops=dict(arrowstyle='-', color='gray',
                                        lw=0.6, shrinkA=0, shrinkB=4))

    ax.set_xlabel('展向位置 (r/R)')
    ax.set_ylabel('攻角 (°)', color='black')

    # ===== 右轴：相对厚度（可由 show_thickness 开关控制） =====
    if show_thickness and span_pos is not None and span_thickness is not None:
        span_pos = np.asarray(span_pos, dtype=float)
        span_thickness = np.asarray(span_thickness, dtype=float)
        order_t = np.argsort(span_pos)
        xs_t = span_pos[order_t]
        ys_t = span_thickness[order_t]
        f_thick = PchipInterpolator(xs_t, ys_t)
        x_dense_t = np.linspace(xs_t.min(), xs_t.max(), 300)
        ax2 = ax.twinx()
        ax2.plot(x_dense_t, f_thick(x_dense_t),
                 '-', color='#ff7f0e', linewidth=1.8, alpha=0.5, label='相对厚度')
        ax2.set_ylabel('相对厚度 (%)', color='#ff7f0e')
        ax2.tick_params(axis='y', labelcolor='#ff7f0e')
        # 右轴图例合并到左轴
        h1, l1 = ax.get_legend_handles_labels()
        h2, l2 = ax2.get_legend_handles_labels()
        ax.legend(h1 + h2, l1 + l2, loc='best')
    else:
        ax.legend(loc='best')

    ax.set_title('失速攻角 / 最大攻角 / 相对厚度 展向分布')
    ax.grid(True, alpha=0.3)
