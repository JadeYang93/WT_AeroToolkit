"""Exporters for standalone shape design data."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import numpy as np
from openpyxl import Workbook, load_workbook

from .models import ShapeDesignOptions, ShapeDesignResult


def _write_table(path: Path, header: list[str], data: np.ndarray) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.append(header)
    for row in np.asarray(data):
        ws.append([float(v) for v in row])
    wb.save(path)


def _append_focus_sheet(path: Path, spans_m: np.ndarray, toSS: np.ndarray, toPS: np.ndarray) -> None:
    """在 trailing_edge_thickness.xlsx 追加一个 'Focus' sheet（Focus 软件接受的尾缘厚度格式）。

    列定义：
      1. span(mm)   = span_m × 1000
      2. flag       全为 1
      3. toSS(mm)   = -原toSS；找到最小值（最负）所在展向位置，
                      从该位置起到叶尖全部填该最小值
      4. toPS(mm)   = 原toPS（保持负值）；找到最小值所在展向位置，
                      从根部到该位置全部填该最小值
    """

    spans_m = np.asarray(spans_m, dtype=float)
    toSS = np.asarray(toSS, dtype=float)
    toPS = np.asarray(toPS, dtype=float)
    span_mm = spans_m * 1000.0

    ss = -toSS
    ps = toPS.copy()
    if ss.size > 0:
        idx_ss = int(np.argmin(ss))
        ss[idx_ss:] = ss[idx_ss]
    if ps.size > 0:
        idx_ps = int(np.argmin(ps))
        ps[: idx_ps + 1] = ps[idx_ps]

    ones = np.ones_like(span_mm)
    table = np.column_stack([span_mm, ones, ss, ps])

    wb = load_workbook(path)
    if "Focus" in wb.sheetnames:
        del wb["Focus"]
    ws = wb.create_sheet("Focus")
    ws.append(["span(mm)", "flag", "toSS(mm)", "toPS(mm)"])
    for row in table:
        ws.append([float(v) for v in row])
    wb.save(path)


def _write_matlab_airfoil_table(path: Path, station_labels: list[str], data: np.ndarray) -> None:
    """Write standard airfoil points like MATLAB points_cloud=[head_1;head_2;num2cell(std_points)]."""

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active

    head_1: list[str] = []
    head_2: list[str] = []
    for label in station_labels:
        head_1.extend([label, label])
        head_2.extend(["x", "y"])

    ws.append(head_1)
    ws.append(head_2)
    for row in np.asarray(data):
        ws.append([float(v) for v in row])
    wb.save(path)


def _matrix_header(prefix: str, labels: tuple[str, ...], n_sections: int) -> list[str]:
    header: list[str] = []
    for idx in range(n_sections):
        for label in labels:
            header.append(f"{prefix}{idx + 1}_{label}")
    return header


def station_label(span_m: float) -> str:
    """Return section label compatible with the correction workflow."""

    value = float(span_m)
    if abs(value - round(value)) < 1e-10:
        return f"R{int(round(value))}"
    text = f"{value:g}".replace(".", "_").replace("-", "m")
    return f"R{text}"


def matlab_station_label(span_m: float) -> str:
    """Return section label like MATLAB strcat('R',num2str(x_span(k)))."""

    value = float(span_m)
    return f"R{value:g}"


def _section_matrix_header(result: ShapeDesignResult, labels: tuple[str, ...]) -> list[str]:
    spans = result.sampled_geo[:, 0]
    station_labels = result.station_labels or [station_label(span) for span in spans]
    header: list[str] = []
    for base in station_labels:
        for label in labels:
            header.append(f"{base}_{label}")
    return header


def write_focus_file(path: str | Path, result: ShapeDesignResult) -> Path:
    """Write Focus.mac using the same public arguments as MATLAB Output_focus."""

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    span = result.sampled_geo[:, 0]
    chord = result.sampled_geo[:, 1]
    twist = result.sampled_geo[:, 2]
    pitchaxis = result.sampled_geo[:, 4]
    preb = result.sampled_geo[:, 5]
    # sweep 默认 None → 全 0；第二阶段从 GEO_for_correction 读
    if result.sweep is None:
        sweep = np.zeros_like(span, dtype=float)
    else:
        sweep = np.asarray(result.sweep, dtype=float)
    std_points = result.standard_points

    with path.open("w", encoding="utf-8", newline="") as f:
        f.write(f'"DEF PARA,         RADIUS={span[-1]:5.2f}"\r\n')
        f.write("COMMENT\r\n\r\nEND OF COMMENT\r\n")
        for idx, sec in enumerate(span):
            focus_x = std_points[:, 2 * idx]
            focus_y = std_points[:, 2 * idx + 1]
            f.write(f"DEF SHAPE R_{sec:<6.2f}  {pitchaxis[idx] / 100:<8.5f}  {0:<8.5f}\r\n")
            for x_val, y_val in zip(focus_x, -focus_y):
                f.write(f" POINTS             {x_val:<9.7f} {y_val:<9.7f}\r\n")
            f.write("END SHAPE/C\r\n")
        for idx, sec in enumerate(span):
            f.write(f"PLACE SHAPE R_{sec:<6.2f}  {sec * 1000:<6.0f}\r\n")
            f.write(f"Rotation            /DG {90 - twist[idx]:<6.2f}\r\n")
            # Translation 第一项是预弯（preb），第二项是后掠（sweep），均以 mm 输出
            f.write(f"Translation         {preb[idx] * 1000:<8.2f}  {sweep[idx] * 1000:<8.2f}\r\n")
            f.write(f"Scale factors      {chord[idx] * 1000:8.2f}  {chord[idx] * 1000:8.2f}\r\n")
            f.write("END PLACE\n")
        f.write("DEF BL-AXIS                OFF\r\n")
        f.write("END DEF BL-AXIS\r\n")
        f.write("COMMENT\r\nEND OF COMMENT\r\n")
    return path


def write_step_points_file(path: str | Path, result: ShapeDesignResult) -> Path:
    """Write CATIA-compatible STEP point cloud using MATLAB Output_CATIA layout."""

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    real_x = result.real_x
    real_y = result.real_y
    real_z = result.real_z
    n_points, n_sections = real_x.shape
    total_points = n_points * n_sections
    relationship_id = 50 + total_points + 1
    geometric_set_id = 50 + total_points + 2
    bounded_rep_id = 50 + total_points + 3
    date_text = datetime.now().strftime("%d-%b-%Y %H:%M:%S")

    with path.open("w", encoding="utf-8", newline="") as f:
        f.write("ISO-10303-21;\n")
        f.write("HEADER;\n\n")
        f.write("/* STEP file generated using Python.*/\n")
        f.write("/* STEP EXPORT developer:GMn&YR */\n")
        f.write("/* Tel: */\n\n")
        f.write("FILE_DESCRIPTION(('Matlab create'),'2;1');\n")
        f.write(f"FILE_NAME('','{date_text}',('none'),('none'),'Python','shape_design','none');\n")
        f.write("FILE_SCHEMA(('CONFIG_CONTROL_DESIGN'));\n")
        f.write("ENDSEC;\n")
        f.write("/* file written by Python shape_design */\n")
        f.write("DATA;\n")
        f.write("#1=APPLICATION_CONTEXT('configuration controlled 3D design of mechanical parts and assemblies') ;\n")
        f.write("#2=MECHANICAL_CONTEXT(' ',#1,'mechanical') ;\n")
        f.write("#3=DESIGN_CONTEXT(' ',#1,'design') ;\n")
        f.write("#4=APPLICATION_PROTOCOL_DEFINITION('international standard','config_control_design',1994,#1) ;\n")
        f.write(r"#5=PRODUCT('922A4 - \X2\526F672C\X0\','','',(#2)) ;" + "\n")
        f.write("#6=PRODUCT_DEFINITION_FORMATION_WITH_SPECIFIED_SOURCE('',' ',#5,.MADE.) ;\n")
        f.write("#7=PRODUCT_CATEGORY('part',$) ;\n")
        f.write("#8=PRODUCT_RELATED_PRODUCT_CATEGORY('detail',$,(#5)) ;\n")
        f.write("#9=PRODUCT_CATEGORY_RELATIONSHIP(' ',' ',#7,#8) ;\n")
        f.write("#10=COORDINATED_UNIVERSAL_TIME_OFFSET(0,0,.AHEAD.) ;\n")
        f.write("#11=CALENDAR_DATE(2022,13,6) ;\n")
        f.write("#12=LOCAL_TIME(14,54,38.,#10) ;\n")
        f.write("#13=DATE_AND_TIME(#11,#12) ;\n")
        f.write("#14=PRODUCT_DEFINITION('design',' ',#6,#3) ;\n")
        f.write("#15=SECURITY_CLASSIFICATION_LEVEL('unclassified') ;\n")
        f.write("#16=SECURITY_CLASSIFICATION(' ',' ',#15) ;\n")
        f.write("#17=DATE_TIME_ROLE('classification_date') ;\n")
        f.write("#18=CC_DESIGN_DATE_AND_TIME_ASSIGNMENT(#13,#17,(#16)) ;\n")
        f.write("#19=APPROVAL_ROLE('APPROVER') ;\n")
        f.write("#20=APPROVAL_STATUS('not_yet_approved') ;\n")
        f.write("#21=APPROVAL(#20,' ') ;\n")
        f.write("#22=PERSON(' ',' ',' ',$,$,$) ;\n")
        f.write("#23=ORGANIZATION(' ',' ',' ') ;\n")
        f.write("#24=PERSONAL_ADDRESS(' ',' ',' ',' ',' ',' ',' ',' ',' ',' ',' ',' ',(#22),' ') ;\n")
        f.write("#25=PERSON_AND_ORGANIZATION(#22,#23) ;\n")
        f.write("#26=PERSON_AND_ORGANIZATION_ROLE('classification_officer') ;\n")
        f.write("#27=CC_DESIGN_PERSON_AND_ORGANIZATION_ASSIGNMENT(#25,#26,(#16)) ;\n")
        f.write("#28=DATE_TIME_ROLE('creation_date') ;\n")
        f.write("#29=CC_DESIGN_DATE_AND_TIME_ASSIGNMENT(#13,#28,(#14)) ;\n")
        f.write("#30=CC_DESIGN_APPROVAL(#21,(#16,#6,#14)) ;\n")
        f.write("#31=APPROVAL_PERSON_ORGANIZATION(#25,#21,#19) ;\n")
        f.write("#32=APPROVAL_DATE_TIME(#13,#21) ;\n")
        f.write("#33=CC_DESIGN_PERSON_AND_ORGANIZATION_ASSIGNMENT(#25,#34,(#6)) ;\n")
        f.write("#34=PERSON_AND_ORGANIZATION_ROLE('design_supplier') ;\n")
        f.write("#35=CC_DESIGN_PERSON_AND_ORGANIZATION_ASSIGNMENT(#25,#36,(#6,#14)) ;\n")
        f.write("#36=PERSON_AND_ORGANIZATION_ROLE('creator') ;\n")
        f.write("#37=CC_DESIGN_PERSON_AND_ORGANIZATION_ASSIGNMENT(#25,#38,(#5)) ;\n")
        f.write("#38=PERSON_AND_ORGANIZATION_ROLE('design_owner') ;\n")
        f.write("#39=CC_DESIGN_SECURITY_CLASSIFICATION(#16,(#6)) ;\n")
        f.write("#40=PRODUCT_DEFINITION_SHAPE(' ',' ',#14) ;\n")
        f.write("#41=(LENGTH_UNIT()NAMED_UNIT(*)SI_UNIT(.MILLI.,.METRE.)) ;\n")
        f.write("#42=(NAMED_UNIT(*)PLANE_ANGLE_UNIT()SI_UNIT($,.RADIAN.)) ;\n")
        f.write("#43=PLANE_ANGLE_MEASURE_WITH_UNIT(PLANE_ANGLE_MEASURE(0.0174532925199),#42) ;\n")
        f.write("#44=(NAMED_UNIT(*)SI_UNIT($,.STERADIAN.)SOLID_ANGLE_UNIT()) ;\n")
        f.write("#45=UNCERTAINTY_MEASURE_WITH_UNIT(LENGTH_MEASURE(0.005),#41,'distance_accuracy_value','CONFUSED CURVE UNCERTAINTY') ;\n")
        f.write("#46=(GEOMETRIC_REPRESENTATION_CONTEXT(3)GLOBAL_UNCERTAINTY_ASSIGNED_CONTEXT((#45))GLOBAL_UNIT_ASSIGNED_CONTEXT((#41,#42,#44))REPRESENTATION_CONTEXT(' ',' ')) ;\n")
        f.write("#47=CARTESIAN_POINT(' ',(0.,0.,0.)) ;\n")
        f.write("#48=AXIS2_PLACEMENT_3D(' ',#47,$,$) ;\n")
        f.write("#49=SHAPE_REPRESENTATION(' ',(#48),#46) ;\n")
        f.write("#50=SHAPE_DEFINITION_REPRESENTATION(#40,#49) ;\n\n\n")

        ident = 1
        for sec in range(n_sections):
            for point in range(n_points):
                f.write(
                    f"#{ident + 50}=CARTESIAN_POINT('Sect{sec + 1}_{point + 1}',"
                    f"({real_x[point, sec]:.6f},{real_y[point, sec]:.6f},{real_z[point, sec]:.6f}));\n"
                )
                ident += 1
        f.write(
            f"#{relationship_id}=SHAPE_REPRESENTATION_RELATIONSHIP(' ',' ',#49,#{bounded_rep_id});\n"
        )
        f.write(f"#{geometric_set_id}=GEOMETRIC_SET('NONE',(\n")
        for point_id in range(51, 51 + total_points):
            suffix = "));\n" if point_id == 50 + total_points else ",\n"
            f.write(f"#{point_id}{suffix}")
        f.write(
            f"#{bounded_rep_id}=GEOMETRICALLY_BOUNDED_SURFACE_SHAPE_REPRESENTATION('NONE',"
            f"(#{geometric_set_id}),#46) ;\n"
        )
        f.write("\n")
        f.write("ENDSEC;\nEND-ISO-10303-21;\n")
    return path


def export_shape_design(
    result: ShapeDesignResult,
    output_dir: str | Path,
    options: ShapeDesignOptions | None = None,
    progress_callback=None,
    layout: str = "split",
) -> dict[str, Path]:
    """Write selected shape design files and return their paths.

    progress_callback: 可选 callable(message: str)，每写完一个文件回调一次。

    layout:
        - ``"split"``（默认，STAGE-3 用）：气动几何写到 ``output_dir/AeroGEO/``，
          CATIA 模型写到 ``output_dir/CATIA/``。
        - ``"flat"``（STAGE-1 用）：所有文件直接写到 ``output_dir/`` 顶层，
          不嵌套 AeroGEO/CATIA 子目录。
    """
    def _cb(msg):
        if progress_callback:
            try:
                progress_callback(msg)
            except Exception:
                pass

    options = options or ShapeDesignOptions()
    output_dir = Path(output_dir)
    if layout == "flat":
        aero_dir = output_dir
        catia_dir = output_dir
    else:
        aero_dir = output_dir / "AeroGEO"
        catia_dir = output_dir / "CATIA"
    written: dict[str, Path] = {}

    if options.export_airfoil_points or options.export_3d_points or options.export_geometry:
        n_sections = result.sampled_geo.shape[0]
        sampled = result.sampled_geo
        lead_edge = sampled[:, 1] - sampled[:, 1] * sampled[:, 4] / 100.0
        tail_edge = -sampled[:, 1] * sampled[:, 4] / 100.0
        real_thick = sampled[:, 1] * sampled[:, 3] / 100.0
        # sweep 默认 None（第一阶段未定义）→ 用 0 数组；第二阶段从 GEO_for_correction 读
        if result.sweep is None:
            sweep_col = np.zeros(n_sections, dtype=float)
        else:
            sweep_col = np.asarray(result.sweep, dtype=float)
        aero_geo = np.column_stack([sampled, sweep_col, lead_edge, tail_edge, real_thick])
        matlab_labels = [matlab_station_label(span) for span in sampled[:, 0]]

    if options.export_airfoil_points:
        path = aero_dir / "standard_airfoil_points.xlsx"
        _write_matlab_airfoil_table(path, matlab_labels, result.standard_points)
        written["standard_airfoil_points"] = path
        _cb(f'[导出] {path.name}（{len(matlab_labels)} 翼型 × {result.standard_points.shape[0] if hasattr(result.standard_points, "shape") else "?"} 点）')

    if options.export_3d_points:
        path = aero_dir / "blade_3d_points.xlsx"
        _write_table(path, _section_matrix_header(result, ("X", "Y", "Z")), result.real_points)
        written["blade_3d_points"] = path
        _cb(f'[导出] {path.name}（3D 点云）')

    if options.export_geometry:
        path = aero_dir / "blade_aero_geometry.xlsx"
        _write_table(
            path,
            ["Span", "Chord", "Twist", "Th%", "PitchAxis", "Prebend", "Sweep", "LE", "TE", "RealThick"],
            aero_geo,
        )
        written["blade_aero_geometry"] = path
        _cb(f'[导出] {path.name}（{aero_geo.shape[0]} 截面）')

    if options.export_tail:
        path = aero_dir / "trailing_edge_thickness.xlsx"
        _write_table(path, ["span(m)", "R_thickness", "thickness(mm)", "toSS(mm)", "toPS(mm)"], result.tail_distribution)
        _append_focus_sheet(
            path,
            spans_m=result.tail_distribution[:, 0],
            toSS=result.tail_distribution[:, 3],
            toPS=result.tail_distribution[:, 4],
        )
        written["trailing_edge_thickness"] = path
        _cb(f'[导出] {path.name}（含 Focus sheet）')

    if options.export_focus:
        written["focus"] = write_focus_file(aero_dir / "Focus.mac", result)
        _cb(f'[导出] Focus.mac')

    if options.export_step_points:
        written["step_points"] = write_step_points_file(catia_dir / "3D_points.stp", result)
        _cb(f'[导出] 3D_points.stp（CATIA）')

    if options.export_bladed:
        raise NotImplementedError("Bladed .prj export is reserved; result.bladed_geo is available as interface data.")

    return written
