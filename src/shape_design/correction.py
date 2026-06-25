"""XFOIL 修正工作流：准备可编辑输入、运行 TGAP/TSET 修正、重建外形结果。

合并自原 correction/ 包的 excel_utils / prepare_inputs / xfoil_correct / rebuild_result，
公共入口为 prepare_correction_inputs / run_airfoil_correction / build_result_from_corrected_files。
"""

from __future__ import annotations

import os
import shutil
import subprocess
import sys
from pathlib import Path
from typing import Iterable

import numpy as np
from openpyxl import Workbook, load_workbook

from .compute import build_result_from_airfoil_points
from .exporters import matlab_station_label, station_label


# ============================================================
# Excel 读写工具（原 excel_utils.py）
# ============================================================

def read_sheet(path: str | Path) -> tuple[list[str], np.ndarray]:
    """Read the first worksheet as a header row plus numeric data."""

    # read_only=True 模式下 openpyxl 会持有底层 zip 文件句柄直到 close()
    # Windows 上不显式 close 会导致文件被锁住、无法删除/覆盖
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        raise ValueError(f"Empty workbook: {path}")
    header = [str(v).strip() if v is not None else "" for v in rows[0]]
    data: list[list[float]] = []
    for row in rows[1:]:
        values = []
        has_value = False
        for value in row[: len(header)]:
            if value is None or value == "":
                values.append(np.nan)
                continue
            try:
                values.append(float(value))
                has_value = True
            except (TypeError, ValueError):
                values.append(np.nan)
        if has_value:
            data.append(values)
    return header, np.asarray(data, dtype=float)


def write_sheet(path: str | Path, header: Iterable[str], data: np.ndarray) -> Path:
    """Write a simple one-header worksheet."""

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.append(list(header))
    for row in np.asarray(data, dtype=float):
        ws.append([None if np.isnan(v) else float(v) for v in row])
    wb.save(path)
    return path


def column(data_header: list[str], data: np.ndarray, *names: str) -> np.ndarray:
    """Return a numeric column by trying several accepted names."""

    lookup = {name.lower(): idx for idx, name in enumerate(data_header)}
    for name in names:
        idx = lookup.get(name.lower())
        if idx is not None:
            return data[:, idx]
    raise KeyError(f"Missing column; tried: {names}")


# ============================================================
# 准备修正输入（原 prepare_inputs.py）
# ============================================================

def _find_input_file(root: Path, filename: str) -> Path:
    """Find an input xlsx under ``root``.

    STAGE-1 输出现在直接平铺在 ``stage1/`` 顶层（不再嵌套 ``AeroGEO/``），
    所以只查 ``root/filename``。为兼容历史输出（重组前跑过一次的目录），
    仍保留 ``root/AeroGEO/filename`` 兜底。
    """
    candidates = [root / filename, root / "AeroGEO" / filename]
    for path in candidates:
        if path.exists():
            return path
    raise FileNotFoundError(f"Cannot find {filename} under {root}")


def _wide_airfoil_by_section(
    standard_header: list[str],
    standard_data: np.ndarray,
    spans: np.ndarray,
) -> tuple[list[str], np.ndarray]:
    header: list[str] = []
    cols: list[np.ndarray] = []
    for idx, span in enumerate(spans):
        label = station_label(span)
        header.extend([f"{label}_x", f"{label}_y"])

        new_x = f"{label}_x"
        new_y = f"{label}_y"
        ml_x = f"{matlab_station_label(span)}_x"
        ml_y = f"{matlab_station_label(span)}_y"
        old_x = f"R{idx + 1}_x"
        old_y = f"R{idx + 1}_y"
        names = {name: pos for pos, name in enumerate(standard_header)}
        if new_x in names and new_y in names:
            x_col = standard_data[:, names[new_x]]
            y_col = standard_data[:, names[new_y]]
        elif ml_x in names and ml_y in names:
            x_col = standard_data[:, names[ml_x]]
            y_col = standard_data[:, names[ml_y]]
        elif old_x in names and old_y in names:
            x_col = standard_data[:, names[old_x]]
            y_col = standard_data[:, names[old_y]]
        else:
            raise KeyError(f"Cannot find airfoil columns for section {span:g}")
        cols.extend([x_col, y_col])
    return header, np.column_stack(cols)


def _read_standard_airfoil_sheet(path: Path) -> tuple[list[str], np.ndarray]:
    """Read either the old one-row header sheet or MATLAB-style two-row airfoil sheet."""

    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb.active
        second_row = [str(v).strip().lower() if v is not None else "" for v in next(ws.iter_rows(min_row=2, max_row=2, values_only=True))]
    finally:
        wb.close()
    if any(name in {"x", "y"} for name in second_row):
        return _read_matlab_airfoil_sheet(path)
    return read_sheet(path)


def _read_matlab_airfoil_sheet(path: Path) -> tuple[list[str], np.ndarray]:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb.active
        head_1 = [str(v).strip() if v is not None else "" for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        head_2 = [str(v).strip() if v is not None else "" for v in next(ws.iter_rows(min_row=2, max_row=2, values_only=True))]
        header: list[str] = []
        for base, axis in zip(head_1, head_2):
            header.append(f"{base}_{axis}" if base and axis else "")

        rows: list[list[float]] = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            values: list[float] = []
            has_value = False
            for value in row[: len(header)]:
                if value is None or value == "":
                    values.append(np.nan)
                else:
                    values.append(float(value))
                    has_value = True
            if has_value:
                rows.append(values)
    finally:
        wb.close()
    return header, np.asarray(rows, dtype=float)


def prepare_correction_inputs(
    initial_output_dir: str | Path,
    correction_input_dir: str | Path,
    sweep: float | np.ndarray = 0.0,
    geo_filename: str = "GEO_for_correction.xlsx",
    airfoil_filename: str = "standard_airfoil_for_correction.xlsx",
    baseline_dir: str | Path | None = None,
) -> dict[str, Path]:
    """Create editable correction input workbooks from the first shape output.

    baseline_dir: ``TEth_baseline.npy`` 的写入目录。默认为 ``correction_input_dir``
    （兼容旧行为）。新版 UI 传入 ``输出/blade_shape/_internal/`` 把基准文件
    从用户可见区分离出去。
    """

    initial_output_dir = Path(initial_output_dir)
    correction_input_dir = Path(correction_input_dir)
    baseline_dir = Path(baseline_dir) if baseline_dir is not None else correction_input_dir

    geo_path = _find_input_file(initial_output_dir, "blade_aero_geometry.xlsx")
    standard_path = _find_input_file(initial_output_dir, "standard_airfoil_points.xlsx")
    tail_path = _find_input_file(initial_output_dir, "trailing_edge_thickness.xlsx")

    geo_header, geo_data = read_sheet(geo_path)
    std_header, std_data = _read_standard_airfoil_sheet(standard_path)
    tail_header, tail_data = read_sheet(tail_path)

    spans = column(geo_header, geo_data, "Span", "section")
    chord = column(geo_header, geo_data, "Chord", "Chod")
    twist = column(geo_header, geo_data, "Twist")
    th_pct = column(geo_header, geo_data, "Th%", "Th")
    pitch = column(geo_header, geo_data, "PitchAxis")
    preb = column(geo_header, geo_data, "Prebend", "Pebend")
    te_mm = column(tail_header, tail_data, "thickness(mm)", "TEth")
    # sweep 优先从 blade_aero_geometry 的 Sweep 列读（与第一阶段输出对齐）；
    # 旧表无该列时退回 sweep 参数（标量或数组），默认 0
    try:
        sweep_col = column(geo_header, geo_data, "Sweep", "sweep")
    except KeyError:
        sweep_col = np.full_like(spans, float(sweep)) if np.isscalar(sweep) else np.asarray(sweep, dtype=float)

    no = np.arange(1, len(spans) + 1, dtype=float)
    correction_geo = np.column_stack([no, spans, chord, twist, th_pct, pitch, preb, sweep_col, te_mm])
    correction_geo_path = correction_input_dir / geo_filename
    write_sheet(
        correction_geo_path,
        ["NO", "section", "Chod", "Twist", "Th", "PitchAxis", "Pebend", "sweep", "TEth"],
        correction_geo,
    )

    # 同时保存一份原始 TEth 基准（te_mm 来自 trailing_edge_thickness.xlsx），
    # 供 apply_te_correction 反复调参时用作「原始值」，避免每次跑覆盖基准导致
    # 后续过渡区计算失真（例如 p4 从 0.7 调到 0.8 时，0.7→0.8 区段已是过渡值，
    # 再用作 te_orig 会让过渡幅度接近 0，看起来「没计算」）。
    baseline_path = baseline_dir / "TEth_baseline.npy"
    baseline_path.parent.mkdir(parents=True, exist_ok=True)
    np.save(baseline_path, np.asarray(te_mm, dtype=float))

    airfoil_header, airfoil_data = _wide_airfoil_by_section(std_header, std_data, spans)
    correction_airfoil_path = correction_input_dir / airfoil_filename
    write_sheet(correction_airfoil_path, airfoil_header, airfoil_data)

    return {
        "geo": correction_geo_path,
        "airfoil": correction_airfoil_path,
        "te_baseline": baseline_path,
    }


# ============================================================
# PCHIP 后缘厚度连续性修正（apply_te_correction 的前置步骤）
# ============================================================

def apply_pchip_te_continuity(
    geo_xlsx: str | Path,
    th_range: tuple[float, float] = (30.0, 50.0),
    exclude_range: tuple[float, float] = (39.0, 41.0),
    baseline_te_path: str | Path | None = None,
    min_fit_points: int = 4,
) -> dict:
    """对 GEO 的 TEth 做 PCHIP 连续性修正，消除指定相对厚度区段的不连续。

    在 ``th_range``（如 30~50%）范围内剔除 ``exclude_range``（如 39~41%）的数据点，
    用剩余点构造 PCHIP 曲线 ``TEth = f(Th%)``，再用 PCHIP 插值替换 exclude 范围
    内的 TEth。``th_range`` 之外、exclude 之外的原值全部保留。

    PCHIP（Piecewise Cubic Hermite Interpolation Polynomial）保单调、不过冲，
    比普通三次样条更适合这种「平滑衔接」场景。

    参数
    ----
    geo_xlsx:
        GEO_for_correction.xlsx 路径（会被原地改写 TEth 列）。
    th_range:
        PCHIP 构造的相对厚度范围（%），默认 (30, 50)。
    exclude_range:
        被剔除/重算的相对厚度范围（%），默认 (39, 41)。必须在 th_range 内。
    baseline_te_path:
        TEth_baseline.npy 路径。提供则**同步更新**该 .npy 的 exclude 段，
        避免后续 apply_te_correction 用未修正的 baseline 覆盖本次修正。
    min_fit_points:
        构造 PCHIP 所需的最少数据点（剔除后）。不足则跳过修正。

    返回
    ----
    dict 包含：
        geo              : 被改写的 GEO 路径
        sections         : 展向位置数组
        th_pct           : 相对厚度数组
        te_before        : 修正前 TEth（来自 baseline 或 GEO）
        te_after         : 修正后 TEth
        modified_indices : 被替换的截面索引数组
        fit_x/fit_y      : PCHIP 构造点（同 Th% 已平均）
        status           : 'ok' / 'skipped' / 'partial'
        reason           : 跳过/部分修正时的说明
    """
    from scipy.interpolate import PchipInterpolator

    th_low, th_high = sorted(th_range)
    ex_low, ex_high = sorted(exclude_range)
    if not (th_low <= ex_low < ex_high <= th_high):
        raise ValueError(
            f"exclude_range {exclude_range} 必须在 th_range {th_range} 内"
        )

    geo_path = Path(geo_xlsx)
    header, data = read_sheet(geo_path)
    sections = column(header, data, "section", "Span")
    th_pct = column(header, data, "Th", "Th%")
    te_current = column(header, data, "TEth", "thickness(mm)")
    te_idx = next(
        i for i, name in enumerate(header)
        if name.strip().lower() in {"teth", "thickness(mm)"}
    )

    # 起点 = baseline（若有且形状匹配），否则用 GEO 当前的 TEth
    te_orig = te_current.copy()
    used_baseline = False
    if baseline_te_path is not None and Path(baseline_te_path).exists():
        try:
            bl = np.asarray(np.load(baseline_te_path), dtype=float)
            if bl.shape == te_current.shape:
                te_orig = bl
                used_baseline = True
        except Exception:
            pass

    in_outer = (th_pct >= th_low) & (th_pct <= th_high)
    in_exclude = (th_pct >= ex_low) & (th_pct <= ex_high)
    fit_mask = in_outer & ~in_exclude
    target_indices = np.where(in_exclude)[0]

    if len(target_indices) == 0:
        return {
            "geo": geo_path, "sections": sections, "th_pct": th_pct,
            "te_before": te_orig.copy(), "te_after": te_orig.copy(),
            "modified_indices": np.array([], dtype=int),
            "fit_x": np.array([]), "fit_y": np.array([]),
            "status": "skipped", "reason": f"Th%∈[{ex_low},{ex_high}] 范围内无截面",
        }

    # 同 Th% 多截面 → 取 TEth 平均（PCHIP 要求 x 严格递增）
    buckets: dict[float, list[float]] = {}
    for x, y in zip(th_pct[fit_mask], te_orig[fit_mask]):
        buckets.setdefault(float(x), []).append(float(y))
    fit_x = np.asarray(sorted(buckets.keys()), dtype=float)
    fit_y = np.asarray([float(np.mean(buckets[x])) for x in fit_x], dtype=float)

    if len(fit_x) < min_fit_points:
        return {
            "geo": geo_path, "sections": sections, "th_pct": th_pct,
            "te_before": te_orig.copy(), "te_after": te_orig.copy(),
            "modified_indices": np.array([], dtype=int),
            "fit_x": fit_x, "fit_y": fit_y,
            "status": "skipped",
            "reason": f"构造点不足：剔除后仅 {len(fit_x)} 个（需 ≥{min_fit_points}）",
        }

    pchip = PchipInterpolator(fit_x, fit_y, extrapolate=False)

    # 逐个评估 target：超出 fit 范围的不外推（标记 partial）
    new_te = te_orig.copy()
    evaluated = []
    out_of_range = []
    for idx in target_indices:
        x_q = float(th_pct[idx])
        y_q = float(pchip(np.array([x_q]))[0])
        if not np.isfinite(y_q):
            out_of_range.append(int(idx))
            continue
        new_te[idx] = y_q
        evaluated.append(int(idx))

    modified_indices = np.asarray(evaluated, dtype=int)

    # 没有任何点成功评估（理论上不应发生，因为 exclude 在 fit 范围内）
    if modified_indices.size == 0:
        return {
            "geo": geo_path, "sections": sections, "th_pct": th_pct,
            "te_before": te_orig.copy(), "te_after": te_orig.copy(),
            "modified_indices": modified_indices,
            "fit_x": fit_x, "fit_y": fit_y,
            "status": "skipped", "reason": "所有目标点超出 PCHIP 构造范围",
        }

    # 写回 GEO
    data[:, te_idx] = new_te
    write_sheet(geo_path, header, data)

    # 同步 baseline 的 exclude 段，避免 apply_te_correction 用老 baseline 覆盖
    if used_baseline and baseline_te_path is not None:
        np.save(baseline_te_path, new_te)

    status = "ok" if not out_of_range else "partial"
    reason = ""
    if out_of_range:
        reason = f"{len(out_of_range)} 个目标点超出构造范围未修正：索引 {out_of_range}"

    return {
        "geo": geo_path, "sections": sections, "th_pct": th_pct,
        "te_before": te_orig.copy(), "te_after": new_te.copy(),
        "modified_indices": modified_indices,
        "fit_x": fit_x, "fit_y": fit_y,
        "status": status, "reason": reason,
    }


# ============================================================
# 后缘厚度（TE）批量修正
# ============================================================

def apply_te_correction(
    geo_xlsx: str | Path,
    corr_start: float,
    corr_thickness: float,
    tip_thickness: float,
    fair_start: float,
    blade_radius: float | None = None,
    tail_table_path: str | Path | None = None,
    baseline_te_path: str | Path | None = None,
) -> dict:
    """按余弦光顺过渡规则批量改写 GEO_for_correction.xlsx 的 TEth 列。

    参数
    ----
    corr_start:
        修正区起始无量纲径向位置 p1（如 0.85）。span/R >= p1 到叶尖前最后截面
        全部使用 ``corr_thickness``。
    corr_thickness:
        修正区目标尾缘厚度 p2（mm）。
    tip_thickness:
        叶尖（最后一个展向位置）目标尾缘厚度 p3（mm）。
    fair_start:
        左侧光顺过渡起始无量纲径向位置 p4（如 0.80）。``fair_start`` 到
        ``corr_start`` 之间使用 0.5*(1-cos(πt)) 余弦平滑，从原始值过渡到 p2。
    blade_radius:
        叶尖半径（m）。None 时自动用 GEO 表 ``section`` 列最大值推断（推荐）。
        只有叶片半径与 GEO 末截面不一致时才需要手动指定。
    tail_table_path:
        可选。第一阶段 ``trailing_edge_thickness.xlsx`` 路径，提供则从中读取真实
        的 toPS / toSS（首尾点 y）作为修正前值；不提供或读取失败则按 0.5/0.5
        拆分。修正后 toPS / toSS 按总厚度比例缩放（假设 PS/SS 比例不变）。
    baseline_te_path:
        可选。原始 TEth 基准文件路径（.npy，由 prepare_correction_inputs 写出）。
        提供且存在时，``te_orig`` 从该基准读，保证反复调 p1/p4 时跳过区与过渡区
        都用原始值作起点；不提供或不存在则退回当前 GEO 的 TEth（兼容老文件，
        但反复调参会出现「基准污染」——见 prepare_correction_inputs 的注释）。

    返回
    ----
    dict 包含：
        geo            : 被原地覆盖的 GEO 文件路径
        sections       : 展向位置数组 (m)
        blade_radius   : 实际使用的叶尖半径 (m)
        t_norm         : span/R 无量纲位置
        te_before      : 修正前 TEth 数组 (mm)
        te_after       : 修正后 TEth 数组 (mm)
        toPS_before    : 修正前压力面尾缘厚度 (mm)
        toPS_after     : 修正后压力面尾缘厚度 (mm)
        toSS_before    : 修正前吸力面尾缘厚度 (mm)
        toSS_after     : 修正后吸力面尾缘厚度 (mm)
    """

    if fair_start >= corr_start:
        raise ValueError(
            f"fair_start ({fair_start}) 必须小于 corr_start ({corr_start})，否则过渡段无意义"
        )

    geo_path = Path(geo_xlsx)
    header, data = read_sheet(geo_path)

    sections = column(header, data, "section", "Span")
    if blade_radius is None:
        blade_radius = float(sections.max())
    te_current = column(header, data, "TEth", "thickness(mm)")
    te_idx = next(
        i for i, name in enumerate(header)
        if name.strip().lower() in {"teth", "thickness(mm)"}
    )
    # 读 Th%（相对厚度）列：第二 tab 按 Th% 视角画对比曲线用
    # 老格式可能没有该列 → 用 NaN 填充，第二 tab 会自动跳过
    try:
        th_pct = column(header, data, "Th", "Th%")
    except KeyError:
        th_pct = np.full_like(sections, np.nan, dtype=float)

    # 优先用原始基准（避免反复调参时基准被污染）；找不到或长度不匹配则退回当前 TEth
    te_orig = te_current.copy()
    used_baseline = False
    if baseline_te_path is not None and Path(baseline_te_path).exists():
        try:
            bl = np.asarray(np.load(baseline_te_path), dtype=float)
            if bl.shape == te_current.shape:
                te_orig = bl
                used_baseline = True
        except Exception:
            pass

    t_norm = sections / float(blade_radius)
    new_te = te_orig.copy()
    n = len(sections)
    last_idx = n - 1

    for idx in range(n):
        t = float(t_norm[idx])
        if t < fair_start:
            continue
        if idx == last_idx or t >= 1.0:
            new_te[idx] = float(tip_thickness)
        elif t < corr_start:
            factor = 0.5 * (1.0 - np.cos(np.pi * (t - fair_start) / (corr_start - fair_start)))
            new_te[idx] = te_orig[idx] + (corr_thickness - te_orig[idx]) * factor
        else:
            new_te[idx] = float(corr_thickness)

    data[:, te_idx] = new_te
    write_sheet(geo_path, header, data)

    # toPS / toSS 修正前：优先用第一阶段 trailing_edge_thickness.xlsx 真实值
    toPS_before = te_orig * 0.5
    toSS_before = te_orig * 0.5
    te_table_ref = te_orig.copy()
    if tail_table_path is not None and Path(tail_table_path).exists():
        try:
            t_header, t_data = read_sheet(tail_table_path)
            toPS_before = column(t_header, t_data, "toPS(mm)", "toPS")
            toSS_before = column(t_header, t_data, "toSS(mm)", "toSS")
            te_table_ref = column(t_header, t_data, "thickness(mm)", "TEth")
        except (KeyError, ValueError):
            pass

    # 修正后按总厚度比例缩放（假设 PS/SS 比例不变，真实比例要等第二阶段 XFOIL 跑完）
    with np.errstate(divide="ignore", invalid="ignore"):
        ratio = np.where(te_table_ref > 1e-9, new_te / te_table_ref, 1.0)
    toPS_after = toPS_before * ratio
    toSS_after = toSS_before * ratio

    return {
        "geo": geo_path,
        "sections": sections,
        "blade_radius": float(blade_radius),
        "t_norm": t_norm,
        "th_pct": th_pct,
        "te_before": te_orig.copy(),
        "te_after": new_te.copy(),
        "toPS_before": toPS_before,
        "toPS_after": toPS_after,
        "toSS_before": toSS_before,
        "toSS_after": toSS_after,
    }


# ============================================================
# XFOIL 修正（原 xfoil_correct.py）
# ============================================================

def _column_pair(header: list[str], label: str) -> tuple[int, int]:
    compact = {name.replace(" ", "_"): idx for idx, name in enumerate(header)}
    candidates = [
        (f"{label}_x", f"{label}_y"),
        (label, f"{label}_1"),
    ]
    for x_name, y_name in candidates:
        if x_name in compact and y_name in compact:
            return compact[x_name], compact[y_name]
    starts = [idx for idx, name in enumerate(header) if name.replace(" ", "_").startswith(label)]
    if len(starts) >= 2:
        return starts[0], starts[1]
    raise KeyError(f"Cannot find x/y columns for {label}")


def _strip_te_10(x: np.ndarray, y: np.ndarray) -> tuple[np.ndarray, np.ndarray]:
    tol = 1e-10
    x2 = np.asarray(x, dtype=float).copy()
    y2 = np.asarray(y, dtype=float).copy()
    valid = ~(np.isnan(x2) | np.isnan(y2))
    x2 = x2[valid]
    y2 = y2[valid]
    if len(x2) and abs(x2[0] - 1.0) < tol and abs(y2[0]) < tol:
        x2 = x2[1:]
        y2 = y2[1:]
    if len(x2) and abs(x2[-1] - 1.0) < tol and abs(y2[-1]) < tol:
        x2 = x2[:-1]
        y2 = y2[:-1]
    if len(x2) >= 2:
        dup = np.r_[False, (np.abs(np.diff(x2)) < tol) & (np.abs(np.diff(y2)) < tol)]
        x2 = x2[~dup]
        y2 = y2[~dup]
    return x2, y2


def _add_te_10(x: np.ndarray, y: np.ndarray) -> tuple[np.ndarray, np.ndarray]:
    tol = 1e-10
    x2 = np.asarray(x, dtype=float)
    y2 = np.asarray(y, dtype=float)
    if not (len(x2) and abs(x2[0] - 1.0) < tol and abs(y2[0]) < tol):
        x2 = np.r_[1.0, x2]
        y2 = np.r_[0.0, y2]
    if not (len(x2) and abs(x2[-1] - 1.0) < tol and abs(y2[-1]) < tol):
        x2 = np.r_[x2, 1.0]
        y2 = np.r_[y2, 0.0]
    return x2, y2


def _write_dat(path: Path, x: np.ndarray, y: np.ndarray) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        f.write("airfoil\n")
        for x_val, y_val in zip(x, y):
            f.write(f"{x_val:.8f}  {y_val:.8f}\n")


def _read_dat(path: Path) -> tuple[np.ndarray, np.ndarray]:
    data = np.loadtxt(path, skiprows=1)
    data = np.atleast_2d(data)
    return data[:, 0], data[:, 1]


def _calc_toc_interp(x: np.ndarray, y: np.ndarray) -> float:
    x, y = _strip_te_10(x, y)
    if len(x) < 10:
        return float("nan")
    ile = int(np.argmin(x))
    xu, yu = x[: ile + 1], y[: ile + 1]
    xl, yl = x[ile:], y[ile:]
    iu = np.argsort(xu)
    il = np.argsort(xl)
    xu, yu = xu[iu], yu[iu]
    xl, yl = xl[il], yl[il]
    xmin = max(float(np.min(xu)), float(np.min(xl)))
    xmax = min(float(np.max(xu)), float(np.max(xl)))
    if xmax <= xmin:
        return float("nan")
    xq = np.linspace(xmin, xmax, 401)
    return float(np.max(np.interp(xq, xu, yu) - np.interp(xq, xl, yl)))


def _refit_keep_x_by_surface_399(
    x_ref: np.ndarray,
    x_corr: np.ndarray,
    y_corr: np.ndarray,
    idx_le: int = 200,
) -> np.ndarray:
    if len(x_ref) != 399 or len(x_corr) != 399 or len(y_corr) != 399:
        raise ValueError("refit requires 399 stripped airfoil points")
    x_ref = x_ref.copy()
    x_corr = x_corr.copy()
    y_corr = y_corr.copy()
    x_ref[0] = x_ref[-1] = 0.99999
    x_corr[0] = x_corr[-1] = 0.99999

    cut = idx_le - 1
    xu, yu = x_corr[: idx_le], y_corr[: idx_le]
    xl, yl = x_corr[cut:], y_corr[cut:]
    iu = np.argsort(xu)
    il = np.argsort(xl)
    y_upper = np.interp(x_ref[:idx_le], xu[iu], yu[iu])
    y_lower = np.interp(x_ref[idx_le:], xl[il], yl[il])
    y_ref = np.r_[y_upper, y_lower]
    y_ref[cut] = 0.0
    return y_ref


def _run_xfoil_batch(
    xfoil_exe: Path,
    workdir: Path,
    tasks: list[dict],
    timeout_per_task: int = 30,
) -> None:
    """一次性把所有 XFOIL 命令拼接后只调用 1 次 XFOIL。

    为什么不用每截面一次 subprocess.run：每次启动 console 子进程都可能触发 Windows
    前台焦点事件，几十个截面会让用户感觉到"系统在不停切换窗口"。batch 模式下
    XFOIL 只启动一次，console 只创建一次（再叠加 CREATE_NO_WINDOW + SW_HIDE），
    彻底消除对其他应用的影响。

    tasks 中每个 dict 需包含：
        in_dat, out_dat, do_tgap, tgap, tail_factor, do_tset, toc_target
    """
    if not tasks:
        return

    commands: list[str] = []
    save_to_out: list[tuple[Path, Path]] = []
    for t in tasks:
        rel_in = os.path.relpath(t["in_dat"], workdir)
        save_dat = workdir / f"{t['out_dat'].stem}_save.dat"
        if save_dat.exists():
            save_dat.unlink()
        cmds = [f"LOAD {rel_in}", "GDES"]
        if t["do_tgap"]:
            cmds.extend(["TGAP", f"{t['tgap']:.8f}", f"{t['tail_factor']:.8f}"])
        if t["do_tset"]:
            cmds.extend(["TSET", f"{t['toc_target']:.8f}"])
        # 两行空串对应回车，用于退出 GDES 子菜单回主菜单；PCOP 把 panel points 拷到 airfoil；SAVE 写出。
        cmds.extend(["", "", "PCOP", f"SAVE {save_dat.name}"])
        commands.extend(cmds)
        save_to_out.append((save_dat, t["out_dat"]))
    # 全部任务结束后才 QUIT 退出 XFOIL（之前是每个截面都 QUIT，导致每次重启进程）
    commands.extend(["QUIT", ""])

    log_path = workdir / "xfoil.out"
    workdir.mkdir(parents=True, exist_ok=True)
    # Windows 下双重保险：CREATE_NO_WINDOW 不创建 console 窗口，STARTUPINFO(SW_HIDE) 进一步隐藏。
    # 其他平台这两个常量都不存在。
    creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0)
    startupinfo = None
    if sys.platform == "win32":
        startupinfo = subprocess.STARTUPINFO()
        startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
        startupinfo.wShowWindow = subprocess.SW_HIDE
    # 总超时按任务数线性扩展，避免大叶片被误杀
    total_timeout = max(60, timeout_per_task * len(tasks))
    with log_path.open("w", encoding="utf-8", errors="replace") as log:
        subprocess.run(
            [str(xfoil_exe)],
            cwd=str(workdir),
            input="\n".join(commands),
            text=True,
            stdout=log,
            stderr=subprocess.STDOUT,
            timeout=total_timeout,
            check=False,
            creationflags=creationflags,
            startupinfo=startupinfo,
        )
    for save_dat, out_dat in save_to_out:
        if save_dat.exists():
            out_dat.parent.mkdir(parents=True, exist_ok=True)
            shutil.copyfile(save_dat, out_dat)
    missing = [str(o) for _, o in save_to_out if not o.exists()]
    if missing:
        tail = log_path.read_text(encoding="utf-8", errors="replace")[-3000:] if log_path.exists() else ""
        raise RuntimeError(f"XFOIL 未生成以下输出：{missing}\n{tail}")


def run_airfoil_correction(
    geo_xlsx: str | Path,
    airfoil_xlsx: str | Path,
    outdir: str | Path,
    xfoil_exe: str | Path | None = None,
    workdir: str | Path | None = None,
    enable_tgap: bool = True,
    enable_tset: bool = True,
    th_threshold_tail: float = 60.0,
    th_threshold_thick: float = 40.0,
    tail_factor: float = 0.15,
    te_unit: str = "mm",
    toc_eps: float = 1e-5,
    timeout: int = 30,
    keep_workdir: bool = False,
    progress_callback=None,
) -> dict[str, Path]:
    """Run the MATLAB-style TGAP/TSET correction and write corrected workbooks.

    progress_callback: 可选 callable(message: str)，每完成一个截面 / 阶段切换时
    被调用一次。调用方据此在 UI 显示截面级进度。
    """
    def _cb(msg):
        if progress_callback:
            try:
                progress_callback(msg)
            except Exception:
                pass

    geo_xlsx = Path(geo_xlsx)
    airfoil_xlsx = Path(airfoil_xlsx)
    outdir = Path(outdir)
    # XFOIL 固定路径：src/_bin/xfoil.exe（工具箱约定，不走 ConfigCenter）
    # __file__ = src/shape_design/correction.py → parent.parent = src/
    src_dir = Path(__file__).resolve().parent.parent
    xfoil_exe = Path(xfoil_exe) if xfoil_exe else src_dir / "_bin" / "xfoil.exe"
    workdir = Path(workdir) if workdir else outdir / "xfoil_work"
    dat_dir = outdir / "dat"
    outdir.mkdir(parents=True, exist_ok=True)

    geo_header, geo_data = read_sheet(geo_xlsx)
    foil_header, foil_data = read_sheet(airfoil_xlsx)
    spans = column(geo_header, geo_data, "section", "Span")
    chord = column(geo_header, geo_data, "Chod", "Chord")
    th_pct = column(geo_header, geo_data, "Th", "Th%")
    te_abs = column(geo_header, geo_data, "TEth", "thickness(mm)")

    corrected = foil_data.copy()
    summary = []
    unit_scale = 1e-3 if te_unit.lower() == "mm" else 1e-6

    # ---------- 第一段：准备 base.dat + 收集 XFOIL 任务 ----------
    prepared: list[dict] = []
    xfoil_tasks: list[dict] = []
    total = len(spans)
    _cb(f'共 {total} 个截面待处理')
    for idx, span in enumerate(spans):
        label = station_label(span)
        col_x, col_y = _column_pair(foil_header, label)
        x0 = foil_data[:, col_x]
        y0 = foil_data[:, col_y]
        valid = ~(np.isnan(x0) | np.isnan(y0))
        x0 = x0[valid]
        y0 = y0[valid]

        te_m = float(te_abs[idx]) * unit_scale
        tgap = te_m / float(chord[idx])
        toc_target = float(th_pct[idx]) / 100.0
        toc_before = _calc_toc_interp(x0, y0)
        x_in, y_in = _strip_te_10(x0, y0)

        base_dat = workdir / f"base_{label}_{idx + 1:03d}.dat"
        out_dat = dat_dir / f"corr_{label}_{idx + 1:03d}.dat"
        _write_dat(base_dat, x_in, y_in)

        do_tgap = enable_tgap and float(th_pct[idx]) < th_threshold_tail
        do_tset = enable_tset and float(th_pct[idx]) > th_threshold_thick and abs(toc_before - toc_target) > toc_eps
        need_xfoil = do_tgap or do_tset
        need_fit = do_tset or idx == len(spans) - 1

        x_ref = x_in.copy()
        if len(x_ref) == 399:
            x_ref[0] = x_ref[-1] = 0.99999

        prepared.append({
            "idx": idx, "span": span, "label": label,
            "col_x": col_x, "col_y": col_y,
            "chord": float(chord[idx]), "te_m": te_m,
            "toc_target": toc_target, "toc_before": toc_before,
            "th_pct": float(th_pct[idx]),
            "x_in": x_in, "y_in": y_in, "x_ref": x_ref,
            "need_xfoil": need_xfoil, "need_fit": need_fit,
            "out_dat": out_dat,
        })
        if need_xfoil:
            xfoil_tasks.append({
                "in_dat": base_dat, "out_dat": out_dat,
                "do_tgap": do_tgap, "tgap": tgap, "tail_factor": tail_factor,
                "do_tset": do_tset, "toc_target": toc_target,
            })
        _cb(f'[1/3 准备] {idx+1}/{total}  span={span:.2f} m  th={float(th_pct[idx]):.1f}%  '
            f'toc {toc_before*100:.1f}→{toc_target*100:.1f}%  '
            f'({"TGAP" if do_tgap else "-"}{"/TSET" if do_tset else ""})')

    # ---------- 一次性批量调用 XFOIL（关键：避免反复启动子进程）----------
    est_min = max(1, len(xfoil_tasks) * timeout // 60)
    _cb(f'[2/3 XFOIL] 启动批量修正：{len(xfoil_tasks)}/{total} 个截面需要 XFOIL（预计 ≤{est_min} min）')
    _run_xfoil_batch(
        xfoil_exe=xfoil_exe,
        workdir=workdir,
        tasks=xfoil_tasks,
        timeout_per_task=timeout,
    )
    _cb(f'[2/3 XFOIL] 子进程退出，开始读结果 + 后处理')

    # ---------- 第二段：读 XFOIL 结果 + 后处理（refit / _add_te_10 / 写表）----------
    for p_idx, p in enumerate(prepared):
        idx = p["idx"]
        span = p["span"]
        col_x, col_y = p["col_x"], p["col_y"]
        chord_i = p["chord"]
        te_m = p["te_m"]
        toc_target = p["toc_target"]
        toc_before = p["toc_before"]
        x_in, y_in = p["x_in"], p["y_in"]
        x_ref = p["x_ref"]
        out_dat = p["out_dat"]

        if p["need_xfoil"]:
            x_corr, y_corr = _strip_te_10(*_read_dat(out_dat))
        else:
            x_corr, y_corr = x_in, y_in

        if len(x_corr) == 399:
            x_corr[0] = x_corr[-1] = 0.99999
        if p["need_fit"]:
            y_corr = _refit_keep_x_by_surface_399(x_ref, x_corr, y_corr)
            x_corr = x_ref

        xc, yc = _add_te_10(x_corr, y_corr)
        _write_dat(out_dat, xc, yc)

        n = min(len(xc), corrected.shape[0])
        corrected[:, col_x] = np.nan
        corrected[:, col_y] = np.nan
        corrected[:n, col_x] = xc[:n]
        corrected[:n, col_y] = yc[:n]

        te_up = abs(yc[1]) * chord_i * 1000.0 if len(yc) > 2 else np.nan
        te_low = abs(yc[-2]) * chord_i * 1000.0 if len(yc) > 2 else np.nan
        toc_after = _calc_toc_interp(xc, yc)
        summary.append(
            [
                span,
                chord_i,
                p["th_pct"],
                toc_target,
                toc_before,
                toc_after,
                te_m * 1000.0,
                te_up,
                te_low,
            ]
        )
        _cb(f'[3/3 修正] {p_idx+1}/{total}  span={span:.2f} m  '
            f'toc {toc_before*100:.2f}→{toc_after*100:.2f}%  '
            f'TE 上下 {te_up:.2f}/{te_low:.2f} mm')

    summary_path = outdir / "TE_thickness_corrected.xlsx"
    corrected_path = outdir / "standard_airfoil_corrected.xlsx"
    write_sheet(
        summary_path,
        [
            "section_m",
            "chord_m",
            "Th_pct",
            "toc_target",
            "toc_before",
            "toc_after",
            "TE_target_mm",
            "TE_up_mm",
            "TE_low_mm",
        ],
        np.asarray(summary, dtype=float),
    )
    write_sheet(corrected_path, foil_header, corrected)
    if not keep_workdir:
        shutil.rmtree(workdir, ignore_errors=True)
    return {"summary": summary_path, "corrected_airfoil": corrected_path, "dat_dir": dat_dir}


# ============================================================
# 重建结果（原 rebuild_result.py）
# ============================================================

def _read_corrected_airfoils(
    header: list[str],
    data: np.ndarray,
    spans: np.ndarray,
    progress_callback=None,
) -> tuple[np.ndarray, np.ndarray, np.ndarray, list[str]]:
    standard_cols: list[np.ndarray] = []
    stripped_y: list[np.ndarray] = []
    labels: list[str] = []
    profile_x: np.ndarray | None = None
    total = len(spans)

    for idx, span in enumerate(spans):
        label = station_label(span)
        col_x, col_y = _column_pair(header, label)
        x = data[:, col_x]
        y = data[:, col_y]
        valid = ~(np.isnan(x) | np.isnan(y))
        x = x[valid]
        y = y[valid]
        standard_cols.extend([x, y])
        xs, ys = _strip_te_10(x, y)
        if profile_x is None:
            profile_x = xs
        elif len(xs) != len(profile_x) or not np.allclose(xs, profile_x, atol=1e-8):
            ys = np.interp(profile_x, xs[::-1] if xs[0] > xs[-1] else xs, ys[::-1] if xs[0] > xs[-1] else ys)
        stripped_y.append(ys)
        labels.append(label)
        if progress_callback:
            try:
                progress_callback(f'[重建] 读修正翼型 {idx+1}/{total}  span={span:.2f} m  pts={len(xs)}')
            except Exception:
                pass

    max_standard = max(len(col) for col in standard_cols)
    standard_points = np.full((max_standard, len(standard_cols)), np.nan, dtype=float)
    for idx, col in enumerate(standard_cols):
        standard_points[: len(col), idx] = col

    return np.asarray(profile_x, dtype=float), np.column_stack(stripped_y), standard_points, labels


def build_result_from_corrected_files(
    geo_xlsx: str | Path,
    corrected_airfoil_xlsx: str | Path,
    write_used_2d: bool = True,
    used_2d_path: str | Path | None = None,
    progress_callback=None,
):
    """Create a ShapeDesignResult from edited GEO data and corrected airfoils.

    progress_callback: 可选 callable(message: str)，读取每个截面时回调一次。
    """
    def _cb(msg):
        if progress_callback:
            try:
                progress_callback(msg)
            except Exception:
                pass

    geo_header, geo_data = read_sheet(geo_xlsx)
    foil_header, foil_data = read_sheet(corrected_airfoil_xlsx)

    spans = column(geo_header, geo_data, "section", "Span")
    chord = column(geo_header, geo_data, "Chod", "Chord")
    twist = column(geo_header, geo_data, "Twist")
    th_pct = column(geo_header, geo_data, "Th", "Th%")
    pitch = column(geo_header, geo_data, "PitchAxis")
    preb = column(geo_header, geo_data, "Pebend", "Prebend")
    try:
        sweep = column(geo_header, geo_data, "sweep", "Sweep")
    except KeyError:
        sweep = np.zeros_like(spans)

    sampled_geo = np.column_stack([spans, chord, twist, th_pct, pitch, preb])
    _cb(f'开始重建：共 {len(spans)} 个截面')
    profile_x, y_foil, standard_points, labels = _read_corrected_airfoils(
        foil_header, foil_data, spans, progress_callback=progress_callback)
    result = build_result_from_airfoil_points(
        sampled_geo=sampled_geo,
        profile_x=profile_x,
        y_foil=y_foil,
        standard_points=standard_points,
        source_geo=sampled_geo,
        bladed_geo=sampled_geo,
        sweep=sweep,
        station_labels=labels,
    )
    _cb('重建 3D 外形完成')

    if write_used_2d:
        used_2d_path = Path(used_2d_path) if used_2d_path else Path(corrected_airfoil_xlsx).with_name("StdAirfoil_used_2D.xlsx")
        header: list[str] = []
        cols: list[np.ndarray] = []
        for idx, label in enumerate(labels):
            header.extend([f"{label}_x", f"{label}_y"])
            cols.extend([profile_x, y_foil[:, idx]])
        write_sheet(used_2d_path, header, np.column_stack(cols))
        _cb(f'写出 2D 翼型表：{used_2d_path.name}')

    return result
